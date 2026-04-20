"""
Flujo combinado: Extracción de PDFs y generación directa de Excel.
Optimizado para uso en bundle con multiprocessing seguro y modo fallback.
"""

import gc
import os
import sys
import time
import traceback
from datetime import datetime
from multiprocessing import Pool, cpu_count
from typing import Any, Dict, List, Optional, Tuple

from utils.logger import Logger

# Agregar el directorio padre al path para importar módulos
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# Importar extractores específicos
from extractores.extractor_afp import extraer_datos_afp
from extractores.extractor_boleta import extraer_datos_boleta
from extractores.extractor_quinta import extraer_datos_quinta

logger = Logger("CorePipeline3_Diagnostico")

# ============================================
# CONFIGURACIÓN DE CARPETAS Y TIPOS
# ============================================

CARPETAS_CONFIG = {
    "1_Boletas": {
        "tipo": "BOLETA",
        "extractor": extraer_datos_boleta,
        "campos_extra": ["fecha_extraida"],
        "nombre_hoja": "1_Boletas",
    },
    "2_Afp": {
        "tipo": "AFP",
        "extractor": extraer_datos_afp,
        "campos_extra": [],
        "nombre_hoja": "2_Afp",
    },
    "3_5ta": {
        "tipo": "QUINTA",
        "extractor": extraer_datos_quinta,
        "campos_extra": [],
        "nombre_hoja": "3_5ta",
    },
    "4_Convocatoria": {
        "tipo": "CONVOCATORIA",
        "extractor": None,
        "campos_extra": [],
        "nombre_hoja": "1ERA",
    },
    "5_CertificadosTrabajo": {
        "tipo": "CERTIFICADO_TRABAJO",
        "extractor": None,
        "campos_extra": [],
        "nombre_hoja": "CTRA",
    },
}

BATCH_SIZE = 100
MAX_SAMPLE_ISSUES = 10


def _iso_now() -> str:
    """Retorna timestamp UTC estable para diagnósticos."""
    return datetime.utcnow().isoformat(timespec="seconds") + "Z"


def contar_pdfs_por_carpeta(ruta_carpeta_trabajo: str) -> Dict[str, int]:
    """Cuenta PDFs por carpeta configurada."""
    conteos: Dict[str, int] = {}

    for nombre_carpeta in CARPETAS_CONFIG:
        ruta_subcarpeta = os.path.join(ruta_carpeta_trabajo, nombre_carpeta)
        if not os.path.isdir(ruta_subcarpeta):
            continue

        conteos[nombre_carpeta] = len(
            [archivo for archivo in os.listdir(ruta_subcarpeta) if archivo.lower().endswith(".pdf")]
        )

    return conteos


def _extraer_numero(filename: str) -> int:
    """Ordena archivos usando el sufijo numérico tras el guion bajo."""
    import re

    match = re.search(r"_(\d+)", filename)
    return int(match.group(1)) if match else 0


def _append_sample_issue(
    sample_issues: List[Dict[str, str]],
    level: str,
    message: str,
    file_name: Optional[str] = None,
) -> None:
    """Agrega una incidencia acotada para reportes sin inflar el payload del pool."""
    if len(sample_issues) >= MAX_SAMPLE_ISSUES:
        return

    issue = {"level": level, "message": message}
    if file_name:
        issue["file"] = file_name
    sample_issues.append(issue)


def _make_folder_summary(
    folder_name: str,
    document_type: str,
    pdf_count: int,
    parquet_path: Optional[str],
    row_count: int,
    elapsed_seconds: float,
    warnings_count: int = 0,
    error_count: int = 0,
    sample_issues: Optional[List[Dict[str, str]]] = None,
    error_message: Optional[str] = None,
    success: bool = True,
) -> Dict[str, Any]:
    """Construye el resumen serializable de una carpeta."""
    return {
        "folder_name": folder_name,
        "document_type": document_type,
        "pdf_count": pdf_count,
        "success": success,
        "parquet_path": parquet_path,
        "row_count": row_count,
        "elapsed_seconds": round(elapsed_seconds, 3),
        "warnings_count": warnings_count,
        "error_count": error_count,
        "sample_issues": sample_issues or [],
        "error_message": error_message,
    }


def _log_folder_summary(summary: Dict[str, Any]) -> None:
    """Registra en el logger un resumen compacto por carpeta."""
    folder_name = summary["folder_name"]
    document_type = summary["document_type"]
    pdf_count = summary["pdf_count"]
    elapsed = int(summary["elapsed_seconds"])

    logger.info(f"📂 Procesando: {folder_name}")
    logger.info(f"   Tipo: {document_type} | PDFs: {pdf_count}")

    for issue in summary.get("sample_issues", [])[:3]:
        issue_file = issue.get("file")
        if issue_file:
            logger.warning(f"   ⚠️ {issue_file}: {issue.get('message', 'Sin detalle')}")
        else:
            logger.warning(f"   ⚠️ {issue.get('message', 'Sin detalle')}")

    if summary.get("parquet_path"):
        logger.info(f"   💾 Parquet: {os.path.basename(summary['parquet_path'])}")

    if summary.get("success"):
        logger.info(f"   ✅ {summary['row_count']} registros en 0m {elapsed}s")
    else:
        logger.error(f"   ❌ Carpeta con fallo: {summary.get('error_message', 'Sin detalle')}")


def _build_tasks(ruta_carpeta_trabajo: str, nombre_base_excel: str) -> List[Tuple[str, str, str, str]]:
    """Prepara argumentos serializables para el procesamiento por carpeta."""
    tareas: List[Tuple[str, str, str, str]] = []

    for nombre_carpeta, config in CARPETAS_CONFIG.items():
        ruta_subcarpeta = os.path.join(ruta_carpeta_trabajo, nombre_carpeta)
        if not os.path.isdir(ruta_subcarpeta):
            logger.warning(f"⚠️ Carpeta '{nombre_carpeta}' no encontrada, omitiendo...")
            continue

        nombre_parquet = f"{nombre_base_excel}_{nombre_carpeta}.parquet"
        ruta_parquet = os.path.join(ruta_carpeta_trabajo, nombre_parquet)
        tareas.append((nombre_carpeta, ruta_subcarpeta, config["tipo"], ruta_parquet))

    return tareas


def procesar_carpeta_mp(args: Tuple[str, str, str, str]) -> Dict[str, Any]:
    """
    Procesa una carpeta completa en un proceso separado (SPAWN-SAFE).
    Retorna un payload compacto para evitar fallos al enviar resultados por pipe.
    """
    nombre_carpeta, ruta_carpeta, tipo_documento, ruta_parquet = args
    inicio = time.time()

    try:
        # Re-importar extractores dentro del subproceso (Windows spawn-safe)
        from extractores.extractor_afp import extraer_datos_afp as extractor_afp
        from extractores.extractor_boleta import extraer_datos_boleta as extractor_boleta
        from extractores.extractor_quinta import extraer_datos_quinta as extractor_quinta

        extractores_map = {
            "BOLETA": extractor_boleta,
            "AFP": extractor_afp,
            "QUINTA": extractor_quinta,
            "CONVOCATORIA": None,
            "CERTIFICADO_TRABAJO": None,
        }

        extractor = extractores_map.get(tipo_documento)
        registros: List[Dict[str, Any]] = []
        sample_issues: List[Dict[str, str]] = []
        warnings_count = 0
        error_count = 0

        archivos_pdf = [archivo for archivo in os.listdir(ruta_carpeta) if archivo.lower().endswith(".pdf")]
        archivos_pdf.sort(key=_extraer_numero)

        if extractor is None:
            for archivo in archivos_pdf:
                registros.append({"archivo": archivo})
        else:
            for idx, archivo in enumerate(archivos_pdf, 1):
                ruta_completa = os.path.join(ruta_carpeta, archivo)
                resultado = extractor(ruta_completa)

                registro = {
                    "archivo_original": archivo,
                    "tipo_documento": tipo_documento,
                    "nombre_extraido": resultado.get("nombre"),
                    "dni_extraido": resultado.get("dni"),
                }

                if resultado.get("fecha"):
                    registro["fecha_extraida"] = resultado["fecha"]

                if not resultado.get("exito", False):
                    warnings_count += 1
                    observaciones = resultado.get("observaciones", "Sin detalle")
                    registro["observaciones"] = observaciones
                    _append_sample_issue(sample_issues, "warning", observaciones, archivo)

                registros.append(registro)

                if idx % BATCH_SIZE == 0:
                    gc.collect()

        if registros:
            import pandas as pd

            df = pd.DataFrame(registros)
            df.to_parquet(ruta_parquet, engine="pyarrow", compression="snappy", index=False)
            del df
        else:
            ruta_parquet = None

        elapsed = time.time() - inicio
        row_count = len(registros)
        del registros
        gc.collect()

        return _make_folder_summary(
            folder_name=nombre_carpeta,
            document_type=tipo_documento,
            pdf_count=len(archivos_pdf),
            parquet_path=ruta_parquet,
            row_count=row_count,
            elapsed_seconds=elapsed,
            warnings_count=warnings_count,
            error_count=error_count,
            sample_issues=sample_issues,
            success=True,
        )

    except Exception as exc:
        elapsed = time.time() - inicio
        error_message = f"{type(exc).__name__}: {exc}"
        sample_issues = [
            {
                "level": "error",
                "message": error_message,
            }
        ]
        return _make_folder_summary(
            folder_name=nombre_carpeta,
            document_type=tipo_documento,
            pdf_count=0,
            parquet_path=None,
            row_count=0,
            elapsed_seconds=elapsed,
            warnings_count=0,
            error_count=1,
            sample_issues=sample_issues,
            error_message=error_message,
            success=False,
        )


def procesar_carpetas_paralelo(
    ruta_carpeta_trabajo: str,
    nombre_base_excel: str,
    progress_callback=None,
    execution_mode: str = "parallel",
) -> Dict[str, Any]:
    """
    Procesa todas las carpetas y devuelve un resumen estructurado.
    """
    tareas = _build_tasks(ruta_carpeta_trabajo, nombre_base_excel)
    pre_run_folder_counts = contar_pdfs_por_carpeta(ruta_carpeta_trabajo)

    result: Dict[str, Any] = {
        "success": False,
        "execution_mode": execution_mode,
        "folder_summaries": [],
        "parquet_paths": {},
        "error_message": None,
        "traceback": None,
        "pre_run_folder_counts": pre_run_folder_counts,
    }

    if not tareas:
        result["error_message"] = "No hay carpetas válidas para procesar"
        logger.error(result["error_message"])
        return result

    logger.info(f"🚀 Iniciando procesamiento {execution_mode} de {len(tareas)} carpetas")

    carpetas_completadas = 0
    total_carpetas = len(tareas)
    inicio_global = time.time()

    def _handle_summary(summary: Dict[str, Any]) -> None:
        nonlocal carpetas_completadas

        carpetas_completadas += 1
        result["folder_summaries"].append(summary)

        if summary.get("parquet_path"):
            result["parquet_paths"][summary["folder_name"]] = summary["parquet_path"]

        _log_folder_summary(summary)

        if progress_callback:
            progress_callback(carpetas_completadas, total_carpetas, summary)

        logger.info(f"📊 Progreso: {carpetas_completadas}/{total_carpetas} carpetas")

    try:
        if execution_mode == "parallel":
            num_workers = min(cpu_count(), total_carpetas)
            with Pool(processes=num_workers) as pool:
                for summary in pool.imap_unordered(procesar_carpeta_mp, tareas):
                    _handle_summary(summary)
        elif execution_mode == "sequential":
            for tarea in tareas:
                summary = procesar_carpeta_mp(tarea)
                _handle_summary(summary)
        else:
            raise ValueError(f"Modo de ejecución no soportado: {execution_mode}")
    except Exception as exc:
        result["error_message"] = f"{type(exc).__name__}: {exc}"
        result["traceback"] = traceback.format_exc()
        logger.error(f"❌ Error procesando carpetas en modo {execution_mode}: {result['error_message']}")
        logger.error(result["traceback"])
        return result

    elapsed_total = int(time.time() - inicio_global)
    logger.info(f"✅ Procesamiento {execution_mode} completado en 0m {elapsed_total}s")

    folder_failures = [summary for summary in result["folder_summaries"] if not summary.get("success", False)]
    if folder_failures:
        result["error_message"] = "Se detectaron fallos en una o más carpetas durante el procesamiento"
        return result

    result["success"] = True
    return result


def generar_excel_streaming(rutas_parquet: Dict[str, str], ruta_excel: str) -> Dict[str, Any]:
    """
    Genera Excel en modo streaming y devuelve un resultado estructurado.
    """
    result: Dict[str, Any] = {
        "success": False,
        "sheet_summaries": [],
        "error_message": None,
        "traceback": None,
        "excel_exists": False,
    }

    try:
        import duckdb
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        inicio = time.time()
        logger.info("📊 Generando Excel en modo streaming...")

        wb = openpyxl.Workbook(write_only=True)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for nombre_carpeta, ruta_parquet in rutas_parquet.items():
            if not ruta_parquet or not os.path.exists(ruta_parquet):
                logger.warning(f"⚠️ Parquet no encontrado: {ruta_parquet}")
                continue

            df = duckdb.read_parquet(ruta_parquet).df()
            df = df.where(df.notna(), None)

            if df.empty:
                logger.warning(f"⚠️ DataFrame vacío para {nombre_carpeta}")
                del df
                gc.collect()
                continue

            config = CARPETAS_CONFIG.get(nombre_carpeta, {})
            nombre_hoja = config.get("nombre_hoja", nombre_carpeta)
            ws = wb.create_sheet(title=nombre_hoja)
            es_listado_simple = len(df.columns) == 1 and "archivo" in df.columns

            header_row = []
            for col_name in df.columns:
                cell = openpyxl.cell.cell.WriteOnlyCell(ws)
                cell.value = "ARCHIVO" if es_listado_simple else col_name.replace("_", " ").upper()
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                header_row.append(cell)

            ws.append(header_row)

            if es_listado_simple:
                ws.column_dimensions["A"].width = 25
            else:
                anchos_estandar = {
                    "A": 25,
                    "B": 20,
                    "C": 20,
                    "D": 15,
                    "E": 15,
                }
                for col_letter, ancho in anchos_estandar.items():
                    ws.column_dimensions[col_letter].width = ancho

            datos_numpy = df.to_numpy().tolist()
            for row in datos_numpy:
                ws.append(row)

            result["sheet_summaries"].append(
                {
                    "folder_name": nombre_carpeta,
                    "sheet_name": nombre_hoja,
                    "row_count": len(df),
                }
            )

            logger.info(f"   ✅ Hoja '{nombre_hoja}' escrita ({len(df)} registros)")

            del df
            del datos_numpy
            gc.collect()

        if not result["sheet_summaries"]:
            result["error_message"] = "No se encontraron datos tabulares para exportar a Excel"
            return result

        wb.save(ruta_excel)
        elapsed = int(time.time() - inicio)
        result["excel_exists"] = os.path.exists(ruta_excel)

        if not result["excel_exists"]:
            result["error_message"] = "El archivo Excel no existe después de guardar"
            return result

        logger.info(f"✅ Excel generado en 0m {elapsed}s")
        result["success"] = True
        return result

    except ImportError as exc:
        result["error_message"] = f"Dependencia faltante: {exc}"
        logger.error(f"⚠️ {result['error_message']}")
        return result
    except Exception as exc:
        result["error_message"] = f"{type(exc).__name__}: {exc}"
        result["traceback"] = traceback.format_exc()
        logger.error(f"❌ Error generando Excel: {result['error_message']}")
        logger.error(result["traceback"])
        return result


def procesar_diagnostico_a_excel(
    ruta_carpeta_trabajo: str,
    ruta_excel_final: str,
    progress_callback=None,
    guardar_json_opcional: bool = False,
    execution_mode: str = "parallel",
    diagnostic_path: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Procesa PDFs y genera Excel, devolviendo un resultado estructurado.
    """
    del guardar_json_opcional

    result: Dict[str, Any] = {
        "success": False,
        "execution_mode": execution_mode,
        "excel_path": ruta_excel_final,
        "diagnostic_path": diagnostic_path,
        "folder_summaries": [],
        "parquet_paths": {},
        "sheet_summaries": [],
        "error_message": None,
        "traceback": None,
        "excel_exists": False,
        "started_at": _iso_now(),
        "finished_at": None,
        "frozen": getattr(sys, "frozen", False),
        "work_folder": ruta_carpeta_trabajo,
        "pre_run_folder_counts": {},
    }

    try:
        logger.info("=" * 60)
        logger.info(f"🚀 DIAGNÓSTICO OPTIMIZADO - MODO {execution_mode.upper()}")
        logger.info("=" * 60)

        if not os.path.isdir(ruta_carpeta_trabajo):
            result["error_message"] = f"❌ Carpeta '{ruta_carpeta_trabajo}' no existe"
            logger.error(result["error_message"])
            return result

        nombre_base_excel = os.path.splitext(os.path.basename(ruta_excel_final))[0]
        result["pre_run_folder_counts"] = contar_pdfs_por_carpeta(ruta_carpeta_trabajo)

        folder_processing = procesar_carpetas_paralelo(
            ruta_carpeta_trabajo=ruta_carpeta_trabajo,
            nombre_base_excel=nombre_base_excel,
            progress_callback=progress_callback,
            execution_mode=execution_mode,
        )

        result["folder_summaries"] = folder_processing.get("folder_summaries", [])
        result["parquet_paths"] = folder_processing.get("parquet_paths", {})

        if not folder_processing.get("success", False):
            result["error_message"] = folder_processing.get("error_message") or "No se completó el procesamiento de carpetas"
            result["traceback"] = folder_processing.get("traceback")
            logger.error(f"❌ {result['error_message']}")
            return result

        if not result["parquet_paths"]:
            result["error_message"] = "❌ No se generaron Parquets válidos para construir el Excel"
            logger.error(result["error_message"])
            return result

        excel_result = generar_excel_streaming(result["parquet_paths"], ruta_excel_final)
        result["sheet_summaries"] = excel_result.get("sheet_summaries", [])
        result["excel_exists"] = excel_result.get("excel_exists", False)

        if not excel_result.get("success", False):
            result["error_message"] = excel_result.get("error_message") or "Error desconocido generando el Excel"
            result["traceback"] = excel_result.get("traceback")
            logger.error(f"❌ {result['error_message']}")
            return result

        if not os.path.exists(ruta_excel_final):
            result["error_message"] = "El archivo Excel no existe después de una generación reportada como exitosa"
            result["excel_exists"] = False
            logger.error(f"❌ {result['error_message']}")
            return result

        result["excel_exists"] = True
        result["success"] = True
        logger.info("✅ Proceso completado")
        logger.info("📁 Archivos .parquet mantenidos en carpeta de trabajo (política empresarial)")
        return result

    except Exception as exc:
        result["error_message"] = f"{type(exc).__name__}: {exc}"
        result["traceback"] = traceback.format_exc()
        logger.error(f"❌ Error en generación de diagnóstico: {result['error_message']}")
        logger.error(result["traceback"])
        return result
    finally:
        result["finished_at"] = _iso_now()


def obtener_ruta_carpeta(ruta: Optional[str] = None) -> Optional[str]:
    """Obtiene la ruta de trabajo desde parámetro o selector gráfico."""
    if ruta and os.path.isdir(ruta):
        return os.path.normpath(ruta)

    try:
        from PySide6.QtWidgets import QApplication, QFileDialog

        app = QApplication.instance()
        if app is None:
            app = QApplication(sys.argv)

        ruta_seleccionada = QFileDialog.getExistingDirectory(
            None,
            "Seleccionar carpeta de trabajo",
            "",
            QFileDialog.Option.ShowDirsOnly,
        )

        return os.path.normpath(ruta_seleccionada) if ruta_seleccionada else None
    except Exception as exc:
        logger.error(f"❌ Error al abrir explorador: {exc}")
        return None


if __name__ == "__main__":
    import multiprocessing

    multiprocessing.freeze_support()

    if getattr(sys, "frozen", False):
        multiprocessing.set_start_method("spawn", force=True)

    ruta_trabajo = obtener_ruta_carpeta()
    if ruta_trabajo:
        timestamp = time.strftime("%d.%m.%Y_%H.%M.%S")
        nombre_excel = f"diagnostico_consolidado_{timestamp}.xlsx"
        ruta_excel = os.path.join(ruta_trabajo, nombre_excel)
        resultado = procesar_diagnostico_a_excel(ruta_trabajo, ruta_excel, progress_callback=None)
        if resultado.get("success"):
            print(f"✅ Excel generado: {ruta_excel}")
        else:
            print(f"❌ Error: {resultado.get('error_message', 'Sin detalle')}")
    else:
        print("⛔ No se seleccionó carpeta. Operación cancelada.")
