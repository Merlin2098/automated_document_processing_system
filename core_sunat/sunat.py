import os
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
import sys
from tkinter import Tk, filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.cell.cell import TYPE_STRING

# Importar el extractor (ajustar ruta según estructura final)
# Agregar el directorio padre al path para poder importar desde extractores
current_dir = os.path.dirname(os.path.abspath(__file__))
parent_dir = os.path.dirname(current_dir)
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

from extractores.extractor_sunat import SunatDocumentExtractor


class SUNATDiagnosticGenerator:
    """Genera diagnóstico JSON de documentos SUNAT en una carpeta."""
    
    def __init__(self, folder_path, max_workers=4):
        """
        Inicializa el generador de diagnóstico.
        
        Args:
            folder_path (str): Ruta de la carpeta con PDFs SUNAT
            max_workers (int): Número de hilos para procesamiento paralelo
        """
        self.folder_path = folder_path
        self.max_workers = max_workers
        self.extractor = SunatDocumentExtractor()
        self.results = []
        self.stats = {
            'total_files': 0,
            'processed': 0,
            'errors': 0,
            'alta': 0,
            'baja': 0,
            'otros': 0,
            'sin_datos': 0
        }

    def process_single_pdf(self, filename):
        """
        Procesa un único archivo PDF y extrae su información.
        
        Args:
            filename (str): Nombre del archivo PDF
            
        Returns:
            dict: Información extraída del documento
        """
        pdf_path = os.path.join(self.folder_path, filename)
        
        try:
            document_number, name, doc_type, fecha = self.extractor.extract_document_info(pdf_path)
            
            # Determinar estado del procesamiento
            if document_number and name:
                status = "OK"
                self.stats['processed'] += 1
                
                # Contabilizar tipo de documento
                if doc_type == 'ALTA-SUNAT':
                    self.stats['alta'] += 1
                elif doc_type == 'BAJA-SUNAT':
                    self.stats['baja'] += 1
                else:
                    self.stats['otros'] += 1
            else:
                status = "SIN_DATOS"
                self.stats['sin_datos'] += 1
            
            return {
                'filename': filename,
                'status': status,
                'document_number': document_number,
                'name': name,
                'doc_type': doc_type,
                'fecha': fecha,
                'error': None
            }
            
        except Exception as e:
            self.stats['errors'] += 1
            return {
                'filename': filename,
                'status': "ERROR",
                'document_number': None,
                'name': None,
                'doc_type': None,
                'fecha': None,
                'error': str(e)
            }

    def scan_folder(self):
        """
        Escanea la carpeta y procesa todos los archivos PDF.
        
        Returns:
            list: Lista de resultados del procesamiento
        """
        # Obtener lista de archivos PDF
        pdf_files = [f for f in os.listdir(self.folder_path) if f.lower().endswith('.pdf')]
        self.stats['total_files'] = len(pdf_files)
        
        if not pdf_files:
            return []
        
        # Procesar archivos en paralelo
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {executor.submit(self.process_single_pdf, f): f for f in pdf_files}
            
            for future in as_completed(futures):
                result = future.result()
                self.results.append(result)
        
        return self.results

    def generate_diagnostic_excel(self):
        """
        Genera el archivo Excel de diagnóstico con timestamp.
        
        Returns:
            str: Ruta del archivo Excel generado
        """
        # Generar nombre del archivo con timestamp
        timestamp = datetime.now().strftime("%d.%m.%y_%H.%M.%S")
        excel_filename = f"diagnostico_sunat_{timestamp}.xlsx"
        excel_path = os.path.join(self.folder_path, excel_filename)
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Diagnóstico SUNAT"
        
        # Configurar encabezados (fila 5)
        headers = ['Filename', 'Name', 'Document Number', 'Fecha']
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Escribir datos desde la fila 6
        row_num = 6
        for result in sorted(self.results, key=lambda x: x['filename']):
            ws.cell(row=row_num, column=1).value = result['filename']
            
            # Limpiar comas del nombre
            clean_name = result['name'].replace(',', '') if result['name'] else ""
            ws.cell(row=row_num, column=2).value = clean_name
            
            # Formatear DNI como texto para conservar ceros a la izquierda
            dni_cell = ws.cell(row=row_num, column=3)
            dni_cell.value = result['document_number'] if result['document_number'] else ""
            dni_cell.data_type = TYPE_STRING
            
            ws.cell(row=row_num, column=4).value = result['fecha'] if result['fecha'] else ""
            row_num += 1
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        
        # Agregar metadata en las primeras filas
        ws['A1'] = 'DIAGNÓSTICO SUNAT'
        ws['A1'].font = Font(size=16, bold=True)
        
        ws['A2'] = 'Fecha de análisis:'
        ws['B2'] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        
        ws['A3'] = 'Carpeta analizada:'
        ws['B3'] = self.folder_path
        
        # Guardar archivo
        wb.save(excel_path)
        
        return excel_path

    def run(self):
        """
        Ejecuta el proceso completo de diagnóstico.
        
        Returns:
            tuple: (excel_path, stats)
        """
        print(f"🚀 Iniciando análisis de documentos SUNAT")
        print(f"📁 Carpeta: {self.folder_path}\n")
        
        # Escanear y procesar
        self.scan_folder()
        
        # Generar Excel
        excel_path = self.generate_diagnostic_excel()
        
        # Mostrar resumen
        self._print_summary()
        
        print(f"\n✅ Diagnóstico guardado en: {excel_path}")
        
        return excel_path, self.stats

    def _print_summary(self):
        """Imprime un resumen del análisis."""
        print("\n" + "="*50)
        print("📊 RESUMEN DEL DIAGNÓSTICO")
        print("="*50)
        print(f"📄 Total archivos: {self.stats['total_files']}")
        print(f"✅ Procesados correctamente: {self.stats['processed']}")
        print(f"   • ALTA: {self.stats['alta']}")
        print(f"   • BAJA: {self.stats['baja']}")
        print(f"   • OTROS: {self.stats['otros']}")
        print(f"⚠️  Sin datos extraíbles: {self.stats['sin_datos']}")
        print(f"❌ Errores: {self.stats['errors']}")
        print("="*50)


def seleccionar_carpeta():
    """
    Abre un diálogo para que el usuario seleccione una carpeta.
    
    Returns:
        str: Ruta de la carpeta seleccionada o None si se canceló
    """
    root = Tk()
    root.withdraw()  # Ocultar la ventana principal
    root.attributes('-topmost', True)  # Mantener el diálogo al frente
    
    folder_path = filedialog.askdirectory(
        title="Selecciona la carpeta con documentos SUNAT"
    )
    
    root.destroy()
    
    return folder_path if folder_path else None


def generar_diagnostico_sunat(folder_path: str, max_workers: int = 4):
    """
    Función de interfaz para generar diagnóstico SUNAT.
    
    Args:
        folder_path (str): Ruta de la carpeta con PDFs
        max_workers (int): Número de hilos paralelos
        
    Returns:
        tuple: (excel_path, stats)
    """
    if not os.path.isdir(folder_path):
        raise ValueError(f"La ruta '{folder_path}' no es una carpeta válida")
    
    generator = SUNATDiagnosticGenerator(folder_path, max_workers)
    return generator.run()


# Punto de entrada para uso standalone
if __name__ == "__main__":
    print("🔍 Selecciona la carpeta con documentos SUNAT...")
    
    # Permitir pasar ruta por argumento o usar selector
    if len(sys.argv) >= 2:
        folder_path = sys.argv[1]
    else:
        folder_path = seleccionar_carpeta()
    
    if not folder_path:
        print("❌ No se seleccionó ninguna carpeta. Proceso cancelado.")
        sys.exit(1)
    
    try:
        excel_path, stats = generar_diagnostico_sunat(folder_path)
        print(f"\n✅ Proceso completado exitosamente")
        sys.exit(0)
    except Exception as e:
        print(f"❌ Error crítico: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)