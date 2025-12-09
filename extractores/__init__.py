"""
Utilidad para convertir JSON a Excel
"""

import json
from typing import List, Dict
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False


def json_to_excel(registros: List[Dict], ruta_excel: str) -> bool:
    """
    Convierte una lista de registros JSON a un archivo Excel.
    
    Args:
        registros: Lista de diccionarios con datos
        ruta_excel: Ruta donde guardar el archivo Excel
        
    Returns:
        True si se guardó correctamente, False en caso contrario
    """
    if not OPENPYXL_DISPONIBLE:
        print("   ⚠️  openpyxl no está instalado. Instalando...")
        try:
            import subprocess
            import sys
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages"])
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            print("   ✅ openpyxl instalado correctamente")
        except Exception as e:
            print(f"   ❌ No se pudo instalar openpyxl: {e}")
            return False
    
    try:
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Diagnóstico"
        
        if not registros:
            print("   ⚠️  No hay registros para convertir")
            return False
        
        # Obtener encabezados (claves del primer registro)
        encabezados = list(registros[0].keys())
        
        # Estilo para encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir encabezados
        for col_idx, encabezado in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = encabezado.replace("_", " ").upper()
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Escribir datos
        for row_idx, registro in enumerate(registros, 2):
            for col_idx, encabezado in enumerate(encabezados, 1):
                valor = registro.get(encabezado, "")
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = valor
                
                # Centrar columnas específicas
                if encabezado in ["tipo_documento", "exito_extraccion", "dni_extraido"]:
                    cell.alignment = Alignment(horizontal="center")
                
                # Colorear según éxito
                if encabezado == "exito_extraccion":
                    if valor:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Ajustar ancho de columnas
        for col_idx, encabezado in enumerate(encabezados, 1):
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            
            # Anchos específicos
            if encabezado == "archivo_original":
                ws.column_dimensions[column_letter].width = 40
            elif encabezado == "observaciones":
                ws.column_dimensions[column_letter].width = 35
            elif encabezado == "nombre_extraido":
                ws.column_dimensions[column_letter].width = 30
            elif encabezado == "tipo_documento":
                ws.column_dimensions[column_letter].width = 15
            elif encabezado == "dni_extraido":
                ws.column_dimensions[column_letter].width = 12
            elif encabezado == "fecha_extraida":
                ws.column_dimensions[column_letter].width = 15
            elif encabezado == "exito_extraccion":
                ws.column_dimensions[column_letter].width = 18
            else:
                ws.column_dimensions[column_letter].width = 20
        
        # Congelar primera fila
        ws.freeze_panes = "A2"
        
        # Guardar archivo
        wb.save(ruta_excel)
        return True
        
    except Exception as e:
        print(f"   ❌ Error al crear Excel: {e}")
        return False


def leer_json(ruta_json: str) -> List[Dict]:
    """
    Lee un archivo JSON y retorna la lista de registros.
    
    Args:
        ruta_json: Ruta del archivo JSON
        
    Returns:
        Lista de diccionarios con datos
    """
    try:
        with open(ruta_json, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"❌ Error al leer JSON: {e}")
        return []